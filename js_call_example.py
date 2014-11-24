#-*- coding: utf-8 -*-

#START - 여기는 javascript에서 call하기 위한 것들
#import sys
#sys.path.append(r'C:\Python27\Lib')
#sys.path.append(r'C:\Python27\Lib\site-packages')
#END

import pymongo
from openpyxl import load_workbook


def stdout(a):
    sys.stdout.write(str(a) + "\\n")


def build(filename,document_name):
    wb = load_workbook(filename)
    ws = wb.active

    #DB 연결에 대한 에러처리 Exception 처리
    collection = db_connection(document_name)

    #각 값이 없을 때 또는 Exception 처리
    key_dict = collection.find_one().keys()
    result = collection.find_one()

    for items in key_dict:
        if items == '_id':
            continue
        #items 의 값 Validation 처리
        ws[items] = result[items]

    wb.save('D:\output.xlsx')
    return 0


def db_connection(document_name):
    try:
        client = pymongo.MongoClient()
        db = client.test
        collection = db[document_name]

        return collection
    except pymongo.errors.ConnectionFailure, e:
        return 1

if __name__ == "__main__":
    filename = sys.argv[1]
    document_name = sys.argv[2]
    stdout(build(filename,document_name))
    sys.exit(0)


# -*- coding: utf-8 -*-
# 오류처리 및 코드를 최대한 최적화 할 것.
import sys

sys.path.append(r'C:\Python27\Lib')
sys.path.append(r'C:\Python27\Lib\site-packages')

import pymongo
from openpyxl import load_workbook
import json
import os
import re

'''
엑셀 셀값에서 다음 값을 가져오는 Function,
이 Function 은 오직 Excel 로만 사용할 것.
이유는 RegExp 방식이 앞이 알파벳임을 가정한 것임.
'''


def get_next(orig_number, row_cnt):

    number_result = map(int, re.findall(r'\d+', orig_number))

    string = re.match(r'[a-zA-Z]+', orig_number).group()

    number = number_result[0] + row_cnt

    result_array = [string, str(number)]

    next_number = ''.join(result_array)

    return next_number

'''
콘솔에 값을 찍어주는 함수
'''


def stdout(a):
    sys.stdout.write(str(a))

'''
갑/을 복사하는 모듈. 서브 값이 특정 값 이상을 넘기면,
페이지를 복사해서 기존 WorkBook 에 넣는 모듈.
'''


def set_sub_sheet(workbook, sub_array, json_info, limit_number, one_page_number):
    print len(sub_array)

    ws = workbook.worksheets[0]
    row_cnt = 0
    page_cnt = 1
    new_page = False
    #TODO >> 여기서부터, 진행, 하나 건너뛰는 문제 해결할 것.
    '''#1. sub array 의 것을 하나씩 빼내면서, '''
    for sub_items in sub_array:
        '''#2. 꺼낸 한 개의 Record 에서 Field 하나 씩 꺼냄. '''
        '''** 행의 수가 첫 번째 페이지를 넘겼을 때, '''
        if row_cnt >= limit_number:
            print u'사이즈가 그 담 페이지로 넘어가겠네'

            if row_cnt > (limit_number+one_page_number*page_cnt):
                new_page = True

            if page_cnt > 1 and new_page:
                workbook.add_sheet(workbook.worksheets[1], page_cnt)
                new_page = False
                page_cnt += 1

            curr_ws = workbook.worksheets[page_cnt]

            for sub_key in sub_items.keys():
                if sub_key == '_id' or sub_key == '__v':
                    continue
                _sub_key = ['_', sub_key]
                next_sub_key = ''.join(_sub_key)

                if next_sub_key in json_info:
                    curr_ws[get_next(json_info[next_sub_key], row_cnt-limit_number)] = sub_items[sub_key]

        else:
            '''행의 수가 첫 번째 페이지를 넘기지 않을 때 또는 첫번째에 해당되는 값은'''
            for sub_key in sub_items.keys():
                if sub_key == '_id' or sub_key == '__v':
                    continue
                #print sub_items[sub_key]
                #설정파일에 array 의 key 가 있으면...
                if sub_key in json_info:
                    ws[get_next(json_info[sub_key], row_cnt)] = sub_items[sub_key]

        row_cnt += 1

    return 1

'''
json 파일을 읽어서 객체를 넘겨주는 함수,
여기서 Exception 또는 Validation 처리도 진행해야 됨.
'''


def get_json_obj(document_name):
    file_path = ''.join([os.getcwd(), '\\config\\', document_name, '_XLS.json'])

    with open(file_path) as jsonObj:
        json_data = json.load(jsonObj)

    return json_data


'''
Company json 파일을 읽어서 객체를 넘겨주는 함수,
여기서 Exception 또는 Validation 처리도 진행해야 됨.
'''


def get_company_obj():
    file_path = ''.join([os.getcwd(), '\\config\\Company.json'])

    with open(file_path) as jsonObj:
        json_data = json.load(jsonObj)

    return json_data


'''
Excel 파일을 만드는 함수,
openpyxl 모듈을 이용해서, template 을 읽은 후에 값을 넣고
결과 path 를 보내준다.
'''


def build(filename, document_name, output_path, document_key):
    print filename, '::', document_name, '::', output_path, '::', document_key
    try:
        # Excel 정보를 가지고 있는 Config 파일 읽어오기.
        json_ = get_json_obj(document_name)

        #Company 정보를 가지고 있는 Config 파일 읽어오기.
        cjson_ = get_company_obj()

        #DB 연결에 대한 에러처리 Exception 처리
        collection = db_connection(document_name.lower())
        #Mongo 에서 데이터를 가져오는 구문.
        result = collection.find_one({"VATKEY": document_key})

        if result is None:
            print u"결과 값이 없습니다.> [%s]" % document_key
            return 1
        else:

            # 엑셀을 읽어온다.
            wb = load_workbook(filename)
            # 워크시트를 가져온다. (기본은 첫번째 시트를..)
            ws = wb.worksheets[0]

            #Mongo 에서 가져온 값을 넣어준다.
            for items in result.keys():

                if items == '_id' or items == '__v':
                    continue

                if items in json_:
                    ws[json_[items]] = result[items]

                if items == 'SUB':
                    sub_result = set_sub_sheet(wb, result[items], json_, 6, 19)
                    if sub_result == 0:
                        print u"SubSet 셋업 오류."
                        return 1

            #기본 Company 정보를 넣어준다.

            for citems in cjson_.keys():

                if citems in json_:
                    ws[json_[citems]] = cjson_[citems]

            #Output Path 에 결과 엑셀을 저장한다.
            opath = ''.join([output_path, document_key, '.xlsx'])

            wb.save(opath)

    except OSError:
        pass
    except:
        print u"의도되지 않은 오류입니다:", sys.exc_info()[0]
        return 1

    return 0


def db_connection(document_name):
    try:
        client = pymongo.MongoClient()
        db = client.test
        collection = db[document_name]

        return collection
    except pymongo.errors.ConnectionFailure, e:
        return 1


# 메인 함수.
if __name__ == "__main__":
    #filename = sys.argv[1]
    #output_path = sys.argv[2]
    #document_name = sys.argv[3]
    #document_key = sys.argv[4]
    #result = build(filename, document_name, output_path, document_key)
    result = build('D:\\v104.xlsx', 'V104', 'D:\\XLSTREAM\\', '20141201V104')
    stdout(result)
    sys.exit(result)


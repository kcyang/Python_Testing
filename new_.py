# -*- coding: utf-8 -*-

import sys, traceback

sys.path.append(r'C:\Python27\Lib')
sys.path.append(r'C:\Python27\Lib\site-packages')

import pymongo
from openpyxl import load_workbook
import json
import os
import re
import copy

'''
엑셀 셀값에서 다음 값을 가져오는 Function,
이 Function 은 오직 Excel 로만 사용할 것.
이유는 RegExp 방식이 앞이 알파벳임을 가정한 것임.
'''

def get_next(orig_number, row_cnt):

    number_result = map(int, re.findall(r'\d+', orig_number))

    string = re.match(r'[a-zA-Z]+', orig_number).group()

    number = number_result[0] + row_cnt

    result_array = [string, str(number)]

    next_number = ''.join(result_array)
    #print 'Next Number is [%s]' % next_number
    return next_number

'''
콘솔에 값을 찍어주는 함수
'''

def stdout(a):
    sys.stdout.write(str(a))

'''
갑/을 복사하는 모듈. 서브 값이 특정 값 이상을 넘기면,
페이지를 복사해서 기존 WorkBook 에 넣는 모듈.
'''

def set_sub_sheet(workbook, sub_array, json_info, limit_number, one_page_number):
    print 'limit_number[%s] one_page_number[%s]' % (limit_number,one_page_number)
    ws = workbook.worksheets[0]
    curr_ws = None
    row_cnt = 0
    page_cnt = 1
    new_page = False

    '''#1. sub array 의 것을 하나씩 빼내면서, '''
    for sub_items in sub_array:
        '''#2. 꺼낸 한 개의 Record 에서 Field 하나 씩 꺼냄. '''
        '''행의 수가 첫 번째 페이지를 넘겼을 때, '''

        '''현재 행이 첫번째 장의 최대수에 도달하면, '''
        if row_cnt >= limit_number:

            '''현재 행의 수가 (첫번째 장 갯수 + 두번째 장의 갯수*시트갯수)보다 크면'''
            if row_cnt >= (limit_number+one_page_number*page_cnt):
                page_cnt += 1
                new_page = True

            '''첫 페이지는 이미 만들어져 있음. 그 것은 그대로 가고, '''
            if page_cnt > 1 and new_page:
                print 'Page count[%s]' % page_cnt
                ws_copy = copy.deepcopy(workbook.worksheets[1])
                workbook.add_sheet(ws_copy, page_cnt)
                curr_ws = workbook.worksheets[page_cnt]
                curr_ws.title = ''.join(str(page_cnt))
                new_page = False

            if page_cnt == 1:
                curr_ws = workbook.worksheets[1]

            print 'Worksheet_name[%s]' % curr_ws

            for sub_key in sub_items.keys():
                if sub_key == '_id' or sub_key == '__v':
                    continue
                _sub_key = ['_', sub_key]
                next_sub_key = ''.join(_sub_key)

                if next_sub_key in json_info:
                    curr_ws[get_next(json_info[next_sub_key], row_cnt-(limit_number+(one_page_number*(page_cnt-1))))] = sub_items[sub_key]

        else:
            '''행의 수가 첫 번째 페이지를 넘기지 않을 때 또는 첫번째에 해당되는 값은'''
            for sub_key in sub_items.keys():
                if sub_key == '_id' or sub_key == '__v':
                    continue
                if sub_key in json_info:
                    ws[get_next(json_info[sub_key], row_cnt)] = sub_items[sub_key]

        row_cnt += 1

    return 1

'''
json 파일을 읽어서 객체를 넘겨주는 함수,
여기서 Exception 또는 Validation 처리도 진행해야 됨.
'''


def get_json_obj(document_name):
    file_path = ''.join([os.getcwd(), '\\config\\', document_name, '_XLS.json'])

    with open(file_path) as jsonObj:
        json_data = json.load(jsonObj)

    return json_data


'''
Company json 파일을 읽어서 객체를 넘겨주는 함수,
여기서 Exception 또는 Validation 처리도 진행해야 됨.
'''

'''
def get_company_obj():
    file_path = ''.join([os.getcwd(), '\\config\\Company.json'])

    with open(file_path) as jsonObj:
        json_data = json.load(jsonObj)

    return json_data
'''


def get_company_obj():
    file_path = ''.join([os.getcwd(), '\\config\\Company.json'])
    try:
        with open(file_path) as jsonObj:
            json_data = json.load(jsonObj)
        return json_data
    except EnvironmentError as e:
        print 'Opps.......', str(e)
    return 0

'''
Excel 파일을 만드는 함수,
openpyxl 모듈을 이용해서, template 을 읽은 후에 값을 넣고
결과 path 를 보내준다.
'''


def build(filename, document_name, output_path, document_key, limit_no, page_no):
    print filename, '::', document_name, '::', output_path, '::', document_key , '::', limit_no, '::', page_no
    try:
        # Excel 정보를 가지고 있는 Config 파일 읽어오기.
        json_ = get_json_obj(document_name)

        #Company 정보를 가지고 있는 Config 파일 읽어오기.
        cjson_ = get_company_obj()

        #DB 연결에 대한 에러처리 Exception 처리
        collection = db_connection(document_name.lower())

        #Mongo 에서 데이터를 가져오는 구문.
        result = collection.find_one({"VATKEY": document_key})

        if result is None:
            print u"결과 값이 없습니다.> [%s]" % document_key
            return 1
        else:

            # 엑셀을 읽어온다.
            wb = load_workbook(filename)
            # 워크시트를 가져온다. (기본은 첫번째 시트를..)
            ws = wb.worksheets[0]

            #Mongo 에서 가져온 값을 넣어준다.
            for items in result.keys():

                if items == '_id' or items == '__v':
                    continue

                if items in json_:
                    ws[json_[items]] = result[items]

                if items == 'SUB':
                    sub_result = set_sub_sheet(wb, result[items], json_, int(limit_no), int(page_no))
                    if sub_result == 0:
                        print u"SubSet 셋업 오류."
                        return 1

            #기본 Company 정보를 넣어준다.

            for citems in cjson_.keys():
                if citems in json_:
                    ws[json_[citems]] = cjson_[citems]

            #일반과세자 신고서의 경우, 고정으로 한번 더 돌려준다.
            if document_name == 'V101':
                ws = wb.worksheets[1]
                for _items in result.keys():
                    if _items == '_id' or _items == '__v':
                        continue
                    ul_key = '!' + _items
                    if ul_key in json_:
                        ws[json_[ul_key]] = result[_items]


            #Output Path 에 결과 엑셀을 저장한다.
            opath = ''.join([output_path, document_key, '.xlsx'])
            
            wb.save(opath)

    except OSError:
        pass
    except IndexError as e:
        print u"인덱스 에러.....", str(e)
    except:
        print u"의도되지 않은 오류입니다:", sys.exc_info()[0]
        return 1

    return 0


def db_connection(document_name):
    try:
        client = pymongo.MongoClient()
        # 현재는 기본으로 BSEVAT를 사용하도록 , 후에 값을 받아서 연결하도록 할 것.(#TODO)
        db = client.BSEVAT
        collection = db[document_name]

        return collection
    except pymongo.errors.ConnectionFailure, e:
        return 1

# 메인 함수.
if __name__ == "__main__":
    #filename = sys.argv[1]
    #output_path = sys.argv[2]
    #document_name = sys.argv[3]
    #document_key = sys.argv[4]
    #limit_no = sys.argv[5]
    #page_no = sys.argv[6]
    #result = build(filename, document_name, output_path, document_key, limit_no, page_no)
    result = build('D:\\V164.xlsx', 'V164', 'D:\\XLSTREAM\\', '20141101V164', 15, 25)
    stdout(result)
    sys.exit(result)

